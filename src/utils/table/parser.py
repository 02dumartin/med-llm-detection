"""
parser.py — master.xlsx 파싱 공통 함수

master 파일 각 시트 구조:
  row N  : "DATASET | conf=X | iou=Y"  ← 블록 시작 헤더
  row N+1: "class", "detection_acc", ... ← 컬럼 헤더
  row N+2~: MA / HE / EX / SE / overall / lesion 데이터
  (빈 행으로 블록 구분)
"""
from pathlib import Path
from openpyxl import load_workbook

from config import TEST_BLOCK_KEYWORDS


def parse_master_sheet(wb, sheet_name: str) -> dict:
    """
    master 파일의 특정 시트를 파싱하여 dataset별, class별 지표 dict 반환.

    Returns:
        {
            "FGART": {
                "MA":      {"detection_acc": 0.27, "AP@0.5": 0.31, ...},
                "HE":      {...},
                "EX":      {...},
                "SE":      {...},
                "overall": {...},
                "lesion":  {...},   # 1cls 시트에만 존재
            },
            "DDR":   {...},
            "E-optha": {...},
            "IDRiD": {...},
            "DIARETDB1": {...},
        }
    """
    if sheet_name not in wb.sheetnames:
        return {}

    ws = wb[sheet_name]
    max_row = ws.max_row
    max_col = ws.max_column

    # 블록 헤더 키워드 → 표준 dataset 이름 역매핑
    # e.g. "eophtha" → "E-optha"
    keyword_to_dataset = {v.lower(): k for k, v in TEST_BLOCK_KEYWORDS.items()}

    result = {}
    current_dataset = None
    col_headers = []   # 현재 블록의 컬럼 헤더 목록

    for r in range(1, max_row + 1):
        row_vals = [ws.cell(r, c).value for c in range(1, max_col + 1)]
        first = row_vals[0]

        # ── 블록 헤더 행 감지: "DATASET | conf=..." ──────────────────
        if isinstance(first, str) and "|" in first:
            # 키워드 매칭 (대소문자 무시)
            matched = None
            for kw, ds in keyword_to_dataset.items():
                if first.lower().startswith(kw):
                    matched = ds
                    break
            if matched:
                current_dataset = matched
                result[current_dataset] = {}
                col_headers = []
            continue

        # ── 컬럼 헤더 행 감지: 첫 셀이 "class" ──────────────────────
        if first == "class":
            col_headers = row_vals
            continue

        # ── 데이터 행 ────────────────────────────────────────────────
        if current_dataset and col_headers and first is not None:
            class_name = str(first).strip()
            row_dict = {}
            for c_idx, header in enumerate(col_headers):
                if header and header != "class":
                    row_dict[header] = row_vals[c_idx]
            result[current_dataset][class_name] = row_dict

    return result


def get_value(parsed: dict, dataset: str, class_name: str, metric: str,
              default=None):
    """parsed dict에서 안전하게 값 꺼내기."""
    try:
        return parsed[dataset][class_name][metric]
    except KeyError:
        return default


def load_master(master_path: Path) -> dict:
    """master.xlsx 전체 로드 → {sheet_name: workbook_sheet} 형태의 openpyxl wb 반환."""
    wb = load_workbook(master_path, data_only=True)
    return wb


def get_row_map(ws, header_row: int = 4) -> dict:
    """
    format 시트에서 (Train, Method) → row 번호 매핑 반환.
    Train 값은 위 행에서 병합 셀로 내려오는 경우 처리.
    """
    row_map = {}
    current_train = None
    for r in range(1, ws.max_row + 1):
        t = ws.cell(r, 1).value
        m = ws.cell(r, 2).value
        if t and t not in ("Train",):
            current_train = str(t).strip()
        if m and m not in ("Method",) and current_train:
            method = str(m).strip()
            row_map[(current_train, method)] = r
    return row_map
