"""
fill_2_main.py — 2_main_acc + 2_main_det 시트 채우기

- 실행: python fill_2_main.py
- 2개 시트를 동시에 처리
- test 축: FGART / DDR / E-optha / IDRiD / DIARETDB1 전부 (고정)
- 소스: overall 행의 값만 사용
"""
import shutil
from pathlib import Path
from openpyxl import load_workbook

from config import (
    MASTER_PATH, FORMAT_PATH, OUTPUT_PATH,
    MODEL_SHEET_MAP_2MAIN as MODEL_SHEET_MAP,
    TEST_COL_START, ACC_COL_OFFSET, DET_COL_OFFSET,
    ALL_TESTSETS,
)
from parser import parse_master_sheet, get_value, get_row_map


def _ensure_output(format_path: Path, output_path: Path):
    """output 파일이 없으면 format 파일 복사해서 생성."""
    if not output_path.exists():
        output_path.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy(format_path, output_path)
        print(f"[init] {output_path} 생성 (format 복사)")
    else:
        print(f"[init] {output_path} 기존 파일에 덮어쓰기")


def fill_2_main():
    _ensure_output(FORMAT_PATH, OUTPUT_PATH)

    wb_master = load_workbook(MASTER_PATH, data_only=True)
    wb_out    = load_workbook(OUTPUT_PATH)

    # ── 시트 정의: (시트명, col_offset_map) ──────────────────────────
    sheet_configs = [
        ("2_main_acc", ACC_COL_OFFSET),
        ("2_main_det", DET_COL_OFFSET),
    ]

    for sheet_name, col_offset_map in sheet_configs:
        ws = wb_out[sheet_name]
        row_map = get_row_map(ws)

        # (Train, Method) 조합 순회
        for (fmt_train, fmt_method), sheet_refs in MODEL_SHEET_MAP.items():
            target_row = row_map.get((fmt_train, fmt_method))
            if target_row is None:
                continue  # format 파일에 해당 행 없음

            # 4cls 시트가 있으면 우선 사용, 없으면 1cls
            master_sheet = sheet_refs.get("4cls") or sheet_refs.get("1cls")
            if master_sheet is None:
                continue

            parsed = parse_master_sheet(wb_master, master_sheet)

            # ── 5개 test dataset 전부 채우기 ─────────────────────────
            for testset in ALL_TESTSETS:
                col_start = TEST_COL_START.get(testset)
                if col_start is None:
                    continue

                for metric, offset in col_offset_map.items():
                    val = get_value(parsed, testset, "overall", metric)
                    if val is not None:
                        ws.cell(target_row, col_start + offset).value = round(float(val), 4)

        print(f"  [{sheet_name}] 완료")

    wb_out.save(OUTPUT_PATH)
    print(f"\n저장 완료: {OUTPUT_PATH}")


if __name__ == "__main__":
    fill_2_main()
