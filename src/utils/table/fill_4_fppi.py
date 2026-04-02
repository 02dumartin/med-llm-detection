"""
fill_4_fppi.py — 4_FPPI_FROC 시트 채우기

설정:
  config.py의 TESTSET_4 변수로 test dataset 선택
  - 단일 선택: "FGART" | "DDR" | "E-optha" | "IDRiD" | "DIARETDB1"
    → 시트명: "4_FPPI_FROC" (format 원본 시트명 그대로)
  - "ALL" 선택:
    → 시트명: "4_FGART", "4_DDR", "4_E-optha", "4_IDRiD", "4_DIARETDB1"
    → output 파일에 5개 시트 자동 생성
"""
import shutil
from pathlib import Path
from openpyxl import load_workbook

from config import (
    MASTER_PATH, FORMAT_PATH, OUTPUT_PATH,
    MODEL_SHEET_MAP_4 as MODEL_SHEET_MAP,
    COL_4, TESTSET_4, ALL_TESTSETS,
)
from parser import parse_master_sheet, get_value, get_row_map

BASE_SHEET = "4_FPPI_FROC"

# master 파일 컬럼명 → format 파일 col 위치
FPPI_METRIC_MAP = {
    "FPPI=0.125": COL_4["FPPI=0.125"],
    "FPPI=0.25":  COL_4["FPPI=0.25"],
    "FPPI=0.5":   COL_4["FPPI=0.5"],
    "FPPI=1.0":   COL_4["FPPI=1.0"],
    "FPPI=2.0":   COL_4["FPPI=2.0"],
    "FPPI=4.0":   COL_4["FPPI=4.0"],
    "FPPI=8.0":   COL_4["FPPI=8.0"],
    "avg_FROC":   COL_4["avg_FROC"],
}


def _ensure_output(format_path: Path, output_path: Path):
    if not output_path.exists():
        output_path.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy(format_path, output_path)
        print(f"[init] {output_path} 생성 (format 복사)")


def _copy_sheet(wb, src_name: str, dst_name: str):
    """같은 workbook 내에서 시트 복사 후 데이터 영역 클리어."""
    if dst_name in wb.sheetnames:
        return wb[dst_name]
    src = wb[src_name]
    dst = wb.copy_worksheet(src)
    dst.title = dst_name
    for row in dst.iter_rows(min_row=3):
        for cell in row:
            if cell.column >= 3:
                cell.value = None
    return dst


def fill_single_testset(ws, wb_master, testset: str):
    """하나의 test dataset에 대해 시트 채우기."""
    row_map = get_row_map(ws)

    for (fmt_train, fmt_method), sheet_refs in MODEL_SHEET_MAP.items():
        target_row = row_map.get((fmt_train, fmt_method))
        if target_row is None:
            continue

        # 4cls 우선, 없으면 1cls
        master_sheet = sheet_refs.get("4cls") or sheet_refs.get("1cls")
        if master_sheet is None:
            continue

        parsed = parse_master_sheet(wb_master, master_sheet)

        # overall 행에서 FPPI + avg_FROC 추출
        for metric, col in FPPI_METRIC_MAP.items():
            val = get_value(parsed, testset, "overall", metric)
            if val is not None:
                ws.cell(target_row, col).value = round(float(val), 4)


def fill_4_fppi():
    _ensure_output(FORMAT_PATH, OUTPUT_PATH)

    wb_master = load_workbook(MASTER_PATH, data_only=True)
    wb_out    = load_workbook(OUTPUT_PATH)

    testsets = ALL_TESTSETS if TESTSET_4 == "ALL" else [TESTSET_4]

    for testset in testsets:
        if TESTSET_4 == "ALL":
            short = testset.replace("-", "").replace(" ", "")
            sheet_name = f"4_{short}"
            ws = _copy_sheet(wb_out, BASE_SHEET, sheet_name)
            print(f"  [{sheet_name}] 시트 생성 (testset={testset})")
        else:
            sheet_name = BASE_SHEET
            ws = wb_out[sheet_name]

        fill_single_testset(ws, wb_master, testset)
        print(f"  [{sheet_name}] testset={testset} 완료")

    wb_out.save(OUTPUT_PATH)
    print(f"\n저장 완료: {OUTPUT_PATH}")


if __name__ == "__main__":
    fill_4_fppi()
