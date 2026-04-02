"""
fill_3_5.py — 3_5_class_acc_det 시트 채우기

설정:
  config.py의 TESTSET_3_5 변수로 test dataset 선택
  - 단일 선택: "FGART" | "DDR" | "E-optha" | "IDRiD" | "DIARETDB1"
    → 시트명: "3_5_class_acc_det" (format 원본 시트명 그대로)
  - "ALL" 선택:
    → 시트명: "3_5_FGART", "3_5_DDR", "3_5_E-optha", "3_5_IDRiD", "3_5_DIARETDB1"
    → output 파일에 5개 시트 자동 생성
"""
import shutil
import copy
from pathlib import Path
from openpyxl import load_workbook

from config import (
    MASTER_PATH, FORMAT_PATH, OUTPUT_PATH,
    MODEL_SHEET_MAP_35 as MODEL_SHEET_MAP,
    COL_35, TESTSET_3_5, ALL_TESTSETS,
)
from parser import parse_master_sheet, get_value, get_row_map

BASE_SHEET = "3_5_class_acc_det"
CLS_LIST   = ["MA", "HE", "EX", "SE"]


def _ensure_output(format_path: Path, output_path: Path):
    if not output_path.exists():
        output_path.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy(format_path, output_path)
        print(f"[init] {output_path} 생성 (format 복사)")


def _copy_sheet(wb, src_name: str, dst_name: str):
    """같은 workbook 내에서 시트 복사."""
    if dst_name in wb.sheetnames:
        return wb[dst_name]
    src = wb[src_name]
    dst = wb.copy_worksheet(src)
    dst.title = dst_name
    # 복사 시 값만 지우기 (헤더/구조 유지, 데이터 영역만 클리어)
    for row in dst.iter_rows(min_row=3):
        for cell in row:
            if cell.column >= 3:  # col A, B (Train/Method) 제외
                cell.value = None
    return dst


def fill_single_testset(ws, wb_master, testset: str):
    """하나의 test dataset에 대해 시트 채우기."""
    row_map = get_row_map(ws)

    for (fmt_train, fmt_method), sheet_refs in MODEL_SHEET_MAP.items():
        target_row = row_map.get((fmt_train, fmt_method))
        if target_row is None:
            continue

        # 4cls 파싱
        parsed_4cls = {}
        if sheet_refs.get("4cls"):
            parsed_4cls = parse_master_sheet(wb_master, sheet_refs["4cls"])

        # 1cls 파싱
        parsed_1cls = {}
        if sheet_refs.get("1cls"):
            parsed_1cls = parse_master_sheet(wb_master, sheet_refs["1cls"])

        # ── AP@0.5 (클래스별) ─────────────────────────────────────
        cls_col_map = {
            "MA": COL_35["AP_MA"],
            "HE": COL_35["AP_HE"],
            "EX": COL_35["AP_EX"],
            "SE": COL_35["AP_SE"],
        }
        for cls, col in cls_col_map.items():
            val = get_value(parsed_4cls, testset, cls, "AP@0.5")
            if val is not None:
                ws.cell(target_row, col).value = round(float(val), 4)

        # ── mAP@0.5 / mAP@0.95 (overall) ─────────────────────────
        for metric, col_key in [("AP@0.5", "mAP05"), ("AP@0.5:0.95", "mAP095")]:
            val = get_value(parsed_4cls, testset, "overall", metric)
            if val is not None:
                ws.cell(target_row, COL_35[col_key]).value = round(float(val), 4)

        # ── Cls. Acc. (클래스별) ──────────────────────────────────
        cls_acc_col_map = {
            "MA": COL_35["cls_acc_MA"],
            "HE": COL_35["cls_acc_HE"],
            "EX": COL_35["cls_acc_EX"],
            "SE": COL_35["cls_acc_SE"],
        }
        for cls, col in cls_acc_col_map.items():
            val = get_value(parsed_4cls, testset, cls, "classification_acc")
            if val is not None:
                ws.cell(target_row, col).value = round(float(val), 4)

        # ── Cls. Acc. Mean (overall) ──────────────────────────────
        val = get_value(parsed_4cls, testset, "overall", "classification_acc")
        if val is not None:
            ws.cell(target_row, COL_35["cls_acc_mean"]).value = round(float(val), 4)

        # ── 1cls mAP@0.5 / mAP@0.95 ─────────────────────────────
        # 1cls 시트의 "lesion" 또는 "overall" 행에서 추출
        for cls_name in ["lesion", "overall"]:
            v05  = get_value(parsed_1cls, testset, cls_name, "AP@0.5")
            v095 = get_value(parsed_1cls, testset, cls_name, "AP@0.5:0.95")
            if v05 is not None:
                ws.cell(target_row, COL_35["1cls_mAP05"]).value  = round(float(v05), 4)
            if v095 is not None:
                ws.cell(target_row, COL_35["1cls_mAP095"]).value = round(float(v095), 4)
            if v05 is not None:
                break  # lesion이 있으면 overall 볼 필요 없음


def fill_3_5():
    _ensure_output(FORMAT_PATH, OUTPUT_PATH)

    wb_master = load_workbook(MASTER_PATH, data_only=True)
    wb_out    = load_workbook(OUTPUT_PATH)

    testsets = ALL_TESTSETS if TESTSET_3_5 == "ALL" else [TESTSET_3_5]

    for testset in testsets:
        if TESTSET_3_5 == "ALL":
            # 시트명: "3_5_FGART", "3_5_DDR" 등
            short = testset.replace("-", "").replace(" ", "")
            sheet_name = f"3_5_{short}"
            ws = _copy_sheet(wb_out, BASE_SHEET, sheet_name)
            print(f"  [{sheet_name}] 시트 생성 (testset={testset})")
        else:
            sheet_name = BASE_SHEET
            ws = wb_out[sheet_name]

        fill_single_testset(ws, wb_master, testset)
        print(f"  [{sheet_name}] testset={testset} 완료")

    wb_out.save(OUTPUT_PATH)
    print(f"\n저장 완료: {OUTPUT_PATH}")


if __name__ == "__main__":
    fill_3_5()
